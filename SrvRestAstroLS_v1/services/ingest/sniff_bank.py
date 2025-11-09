# SrvRestAstroLS_v1/services/ingest/sniff_bank.py
from __future__ import annotations
import re
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Optional, Any, List, Tuple

# ===== Dependencias =====
try:
    import pandas as pd  # type: ignore
except Exception:
    pd = None

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:
    load_workbook = None

# ===== Config / Mapeos =====
ACCOUNT_MAP = {
    "3-111-0100026005-5": {"bank": "ciudad", "display": "Banco Ciudad - CC $"},
    "100-393300535-000": {"bank": "patagonia", "display": "Banco Patagonia - CC $"},
    "163-0-015508/3":    {"bank": "santander", "display": "Banco Santander - CC $"},
}

PREFERRED_GL_SHEET_NAMES = ["archivo contable", "contable", "resumen"]  # intentos por nombre de hoja

BANK_HINTS = [
    ("BANCO CIUDAD", "ciudad"),
    ("BANCO SANTANDER", "santander"),
    ("SANTANDER", "santander"),
    ("BANCO PATAGONIA", "patagonia"),
    ("PATAGONIA", "patagonia"),
]

RE_PERIOD  = re.compile(r"(0?[1-9]|[12][0-9]|3[01])[/\.-](0?[1-9]|1[0-2])[/\.-]([12]\d{3})")
RE_ACCOUNT_ON_LINE = re.compile(r"(?:CC\s*\$|C\/C|\bCTA\.?\s*CTE\b|CUENTA\s*CORRIENTE)[^0-9A-Za-z]*([0-9][0-9\-\./ ]{6,}[0-9])")
RE_PILAGA_ACCOUNT = re.compile(r"\b(\d{3,6}/\d)\b")  # ej: 26005/5

TABLE_HEADER_HINTS = (
    "FECHA DOCUMENTO PRINCIPAL", "DOC. COBRO", "CONTENEDOR", "DETALLE",
    "BENEFICIARIO", "INGRESOS", "EGRESOS", "ACUMULADO",
)
EXCLUDE_TEXT_HINTS = ("SALDO INICIAL", "SALDO FINAL")
PILAGA_HEADER_KEYWORDS = ("RESUMEN CUENTA BANCARIA",)

# ===== Safe wrapper pública =====
def sniff_file(path: Path | str, filename_hint: Optional[str] = None) -> dict:
    """Entry point seguro: nunca levanta excepción."""
    p = Path(path)
    try:
        return sniff_path(p, filename_hint or p.name)
    except Exception as e:
        import traceback
        print("[sniff_file] ERROR:", type(e).__name__, str(e), flush=True)
        print(traceback.format_exc(limit=8), flush=True)
        return {
            "kind": "unknown",
            "detected": {"error": f"{type(e).__name__}: {e}"},
            "table": {"columns": [], "sample": []},
            "suggest": {},
            "needs": {"bank": True, "account_id": False, "period_range": True},
        }

def sniff_path(path: Path, filename_hint: Optional[str] = None) -> dict:
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        return sniff_excel(path, filename_hint or path.name)
    if suffix == ".csv":
        return sniff_csv(path)
    return {"kind": "unknown", "detected": {}, "table": {"columns": [], "sample": []}, "suggest": {}, "needs": {"bank": True, "account_id": False, "period_range": True}}

# ===== Excel principal =====
def sniff_excel(path: Path, filename_hint: Optional[str]) -> dict:
    # Vista previa (para UI)
    cols_preview, rows_preview, min_date_tab, max_date_tab = read_table_preview(path)

    # Header comprimido
    grid = read_excel_header_grid(path, max_rows=20, max_cols=12)
    raw_header_lines = header_lines_from_grid(grid, limit=8)
    compact_header_lines = compact_header(raw_header_lines)
    header_excerpt = "\n".join(compact_header_lines)

    # Nombre de la primera hoja
    first_sheet_name = read_first_sheet_name(path)

    # 1) PILAGA primero (prioridad)
    if looks_like_pilaga(header_excerpt, grid, first_sheet_name, cols_preview):
        kind = "gl"
    else:
        # 2) Extracto si no parece PILAGA
        if header_has_bank_extract_fields(header_excerpt, grid) or columns_look_like_bank(cols_preview):
            kind = "bank_movements"
        else:
            kind = "unknown"

    # Cuenta / Banco
    account_core_dv = header_extract_account(grid)
    if not account_core_dv and kind == "gl":
        account_core_dv = pilaga_extract_account(grid)

    # Fechas por header o por muestra
    header_from, header_to = header_extract_period(grid)
    period_from = header_from or min_date_tab
    period_to   = header_to   or max_date_tab

    # === FAST PATH para CONTABLE (PILAGA) con pandas ===
    if kind == "gl":
        fmin, fmax = fast_pilaga_period_pandas(path)
        if fmin and fmax:
            period_from, period_to = fmin, fmax
        else:
            ws_min, ws_max = scan_worksheet_dates(path)
            period_from = period_from or ws_min
            period_to   = period_to   or ws_max
    else:
        # Extractos: si faltan, escaneo general
        if not (period_from and period_to):
            ws_min, ws_max = scan_worksheet_dates(path)
            period_from = period_from or ws_min
            period_to   = period_to   or ws_max

        # Re-chequeo: si por nombre/columnas es PILAGA, forzamos gl
        if looks_like_pilaga(header_excerpt, grid, first_sheet_name, cols_preview):
            kind = "gl"
            fmin, fmax = fast_pilaga_period_pandas(path)
            if fmin and fmax:
                period_from, period_to = fmin, fmax

    # Banco por mapeo o por texto de header / filename
    bank = None
    account_full: Optional[str] = None

    if account_core_dv and account_core_dv in ACCOUNT_MAP:
        bank = ACCOUNT_MAP[account_core_dv]["bank"]
        account_full = account_core_dv

    if kind == "gl" and not bank and account_core_dv and RE_PILAGA_ACCOUNT.fullmatch(str(account_core_dv)):
        mapped_full = map_short_account_to_full(account_core_dv)
        if mapped_full:
            account_full = mapped_full
            bank = ACCOUNT_MAP[mapped_full]["bank"]

    if not bank and header_excerpt:
        txt = header_excerpt.upper()
        for key, label in BANK_HINTS:
            if key in txt:
                bank = label
                break

    if not bank and filename_hint and kind == "bank_movements":
        low = filename_hint.lower()
        if "ciudad" in low: bank = "ciudad"
        elif "patagonia" in low: bank = "patagonia"
        elif "santander" in low: bank = "santander"

    out = {
        "kind": kind if kind != "unknown" else ("gl" if account_core_dv and RE_PILAGA_ACCOUNT.fullmatch(str(account_core_dv)) else kind),
        "detected": {
            "bank": bank,
            "account_core_dv": account_core_dv,
            "account_full": account_full,
            "header_excerpt": header_excerpt if header_excerpt else None,
            "period_from": period_from,
            "period_to": period_to,
        },
        "table": {"columns": cols_preview, "sample": rows_preview},
        "suggest": {"period_from": period_from, "period_to": period_to},
        "needs": {
            "bank": bank is None,
            "account_id": False,
            "period_range": not (period_from and period_to),
        },
    }
    return out

# ===== Heurísticas de tipo =====
def header_has_bank_extract_fields(header_excerpt: str | None, grid: list[list[str]]) -> bool:
    if not header_excerpt:
        return False
    up = header_excerpt.upper()
    hit_keywords = [
        "EXTRACTO DE CUENTA",
        "TIPO Y NRO. DE CUENTA",
        "DENOMINACIÓN",
        "FECHA DESDE",
        "FECHA HASTA",
    ]
    if any(k in up for k in hit_keywords):
        return True
    labels = { (row[0] or "").strip().upper() for row in (grid or []) if row }
    if {"TIPO Y NRO. DE CUENTA", "DENOMINACIÓN"} & labels:
        return True
    return False

def looks_like_pilaga(header_excerpt: str | None, grid: list[list[str]], first_sheet_name: Optional[str], cols_preview: list[str]) -> bool:
    # 1) Palabras clave en header
    up = (header_excerpt or "").upper()
    if any(k in up for k in PILAGA_HEADER_KEYWORDS):
        return True

    # 2) Cuenta corta (###/d) en primeras filas Y sin “CC $ / Tipo y Nro”
    for row in (grid or []):
        line = " ".join([c for c in row if c])
        up_line = line.upper()
        if RE_PILAGA_ACCOUNT.search(line) and "CC $" not in up_line and "TIPO Y NRO" not in up_line:
            return True

    # 3) Nombre de hoja típico
    if first_sheet_name and first_sheet_name.lower() in PREFERRED_GL_SHEET_NAMES:
        return True

    # 4) Columnas típicas PILAGA
    if columns_look_like_pilaga(cols_preview):
        return True

    return False

def columns_look_like_pilaga(cols: list[str]) -> bool:
    up = [c.strip().upper() for c in (cols or [])]
    has_fecha = any(c == "FECHA" or c.startswith("FECHA") for c in up)
    has_ing = any("INGRESO" in c for c in up)
    has_egr = any("EGRESO" in c for c in up)
    has_acu = any("ACUMULADO" in c for c in up)
    has_doc = any("DOCUMENTO" in c or "DOC." in c for c in up)
    has_ben = any("BENEFICIARIO" in c for c in up)
    score_iea = sum([has_ing, has_egr, has_acu])
    return has_fecha and score_iea >= 2 and (has_doc or has_ben)

def columns_look_like_bank(cols: list[str]) -> bool:
    up = [c.strip().upper() for c in (cols or [])]
    if columns_look_like_pilaga(cols):  # no confundir
        return False
    has_fecha = any(c == "FECHA" or c.startswith("FECHA") for c in up)
    has_desc  = any("DESCRIP" in c or "DETALLE" in c or "CONCEPTO" in c for c in up)
    money_any = any(any(h in c for h in ("DEBE", "HABER", "DÉBITO", "CRÉDITO", "IMPORTE", "MONTO")) for c in up)
    saldo_any = any("SALDO" in c for c in up)
    return has_fecha and (has_desc or money_any or saldo_any)

# ===== Header parsing & helpers =====
def read_excel_header_grid(path: Path, max_rows: int = 20, max_cols: int = 12) -> list[list[str]]:
    if not load_workbook:
        return []
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        grid: list[list[str]] = []
        for r in ws.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=max_cols, values_only=True):
            row = [(str(c).strip() if c not in (None, "") else "") for c in r]
            grid.append(row)
        wb.close()
        return grid
    except Exception:
        return []

def read_first_sheet_name(path: Path) -> Optional[str]:
    if not load_workbook:
        return None
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        name = wb.worksheets[0].title
        wb.close()
        return name
    except Exception:
        return None

def header_lines_from_grid(grid: list[list[str]], limit: int = 8) -> list[str]:
    lines: list[str] = []
    for r in grid[:limit]:
        line = " ".join([c for c in r if c])
        if not line:
            continue
        if looks_like_table_header(line):
            break
        lines.append(line)
        if len(lines) >= 3:
            break
    return lines

def compact_header(lines: list[str]) -> list[str]:
    out: list[str] = []
    for ln in lines:
        if looks_like_table_header(ln):
            continue
        if len(ln) > 120:
            continue
        out.append(ln)
        if len(out) >= 3:
            break
    return out or (lines[:1] if lines else [])

def looks_like_table_header(line: str) -> bool:
    up = line.upper()
    if any(k in up for k in TABLE_HEADER_HINTS):
        return True
    if len([w for w in up.split() if w]) >= 10:
        return True
    return False

def header_find_value(grid: list[list[str]], label: str) -> Optional[str]:
    lab = label.lower()
    for row in grid:
        if not row:
            continue
        if (row[0] or "").strip().lower() == lab:
            for cell in row[1:]:
                if cell:
                    return cell
            return ""
    return None

def header_extract_account(grid: list[list[str]]) -> Optional[str]:
    line = header_find_value(grid, "Tipo y Nro. de Cuenta")
    if line:
        m = RE_ACCOUNT_ON_LINE.search(f"Tipo y Nro. de Cuenta {line}")
        if m:
            return m.group(1).strip()
        return line.strip()
    text = "\n".join([" ".join([c for c in r if c]) for r in grid[:8]])
    m2 = RE_ACCOUNT_ON_LINE.search(text)
    return m2.group(1).strip() if m2 else None

def pilaga_extract_account(grid: list[list[str]]) -> Optional[str]:
    for row in (grid or []):
        line = " ".join([c for c in row if c])
        m = RE_PILAGA_ACCOUNT.search(line)
        if m:
            return m.group(1)
    return None

def header_extract_period(grid: list[list[str]]) -> tuple[Optional[str], Optional[str]]:
    from_val = header_find_value(grid, "Fecha desde")
    to_val   = header_find_value(grid, "Fecha hasta")
    return parse_dmy(from_val), parse_dmy(to_val)

# ===== Limpieza y parse de fechas =====
def clean_date_text(s: Optional[str]) -> Optional[str]:
    if s is None:
        return None
    s = s.replace("\ufeff", "").replace("\xa0", " ").strip()
    s = re.sub(r"^[^\d]+", "", s)
    return s or None

def parse_dmy(s: Optional[str]) -> Optional[str]:
    s = clean_date_text(s)
    if not s:
        return None
    m = RE_PERIOD.search(s)
    if m:
        dd, mm, yyyy = m.group(1), m.group(2), m.group(3)
        try:
            d = date(int(yyyy), int(mm), int(dd))
            return d.isoformat()
        except Exception:
            pass
    try:
        if pd is not None:
            d = pd.to_datetime(s, dayfirst=True, errors="coerce")
            if pd.notna(d):
                if hasattr(d, "iloc"):
                    d = d.iloc[0]
                return d.date().isoformat()
    except Exception:
        pass
    return None

# ===== FAST PATH PILAGA con pandas =====
def fast_pilaga_period_pandas(path: Path) -> tuple[Optional[str], Optional[str]]:
    if pd is None:
        return None, None
    try:
        xls = pd.ExcelFile(str(path), engine="openpyxl")
        # hoja preferida
        sheet = None
        low_names = [n.lower() for n in xls.sheet_names]
        for pref in PREFERRED_GL_SHEET_NAMES:
            if pref in low_names:
                sheet = xls.sheet_names[low_names.index(pref)]
                break
        if sheet is None:
            sheet = xls.sheet_names[0]

        df = pd.read_excel(xls, sheet_name=sheet, header=None)
        if df.shape[1] == 0:
            return None, None

        s = df.iloc[:, 0].astype(str)

        # Detectar fila 'Fecha' (si existe), sino saltar primeras 2
        idx_fecha = s.str.strip().str.lower().eq("fecha")
        start = (int(idx_fecha[idx_fecha].index[0]) + 1) if idx_fecha.any() else 2
        s = s.iloc[start:]

        # Limpiar/filtrar ruido
        s = s[~s.str.upper().str.contains("|".join(EXCLUDE_TEXT_HINTS), na=False)]
        s = s.str.replace(r"^[^\d]+", "", regex=True)

        dt = pd.to_datetime(s, dayfirst=True, errors="coerce").dropna()
        if dt.empty:
            return None, None

        return dt.min().date().isoformat(), dt.max().date().isoformat()
    except Exception:
        return None, None

# ===== Fallback: escaneo general =====
def scan_worksheet_dates(path: Path, max_rows: int = 30000) -> tuple[Optional[str], Optional[str]]:
    if not load_workbook:
        return None, None
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        dmin: Optional[date] = None
        dmax: Optional[date] = None

        for row in ws.iter_rows(min_row=1, max_row=max_rows, values_only=True):
            row_text = " ".join([str(c) for c in row if c]).upper()
            if any(h in row_text for h in EXCLUDE_TEXT_HINTS):
                continue
            for cell in row:
                dt = _as_date(cell)
                if dt:
                    dmin = dt if (dmin is None or dt < dmin) else dmin
                    dmax = dt if (dmax is None or dt > dmax) else dmax

        wb.close()
        return (dmin.isoformat() if dmin else None, dmax.isoformat() if dmax else None)
    except Exception:
        return None, None

def _as_date(val: Any) -> Optional[date]:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        if isinstance(val, float) and not float(val).is_integer():
            return None
        if 20000 <= int(val) <= 80000:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=int(val))).date()
        return None
    if val not in (None, ""):
        iso = parse_dmy(str(val))
        if iso:
            return datetime.fromisoformat(iso).date()
    return None

# ===== Preview tabular =====
def read_table_preview(path: Path) -> tuple[list[str], list[list[Any]], Optional[str], Optional[str]]:
    if pd is None:
        return [], [], None, None
    try:
        df = pd.read_excel(str(path), engine="openpyxl", nrows=120)
        cols = [str(c) for c in df.columns.tolist()]
        sample = df.head(10).fillna("").astype(object).values.tolist()
        min_d, max_d = try_parse_dates_in_df(df)
        return cols, sample, min_d, max_d
    except Exception:
        return [], [], None, None

def try_parse_dates_in_df(df) -> tuple[Optional[str], Optional[str]]:
    if df is None or df.empty:
        return None, None
    date_col = None
    for c in df.columns:
        s = df[c]
        if str(s.dtype).startswith("datetime"):
            date_col = c; break
        if s.dtype == "object" and pd is not None:
            try:
                s2 = s.astype(str).map(clean_date_text)
                parsed = pd.to_datetime(s2, errors="coerce", dayfirst=True, infer_datetime_format=True)
                if parsed.notna().sum() >= max(3, int(len(parsed) * 0.1)):
                    df[c] = parsed
                    date_col = c
                    break
            except Exception:
                pass
    if not date_col:
        return None, None
    try:
        s = pd.to_datetime(df[date_col], errors="coerce").dropna()
        if s.empty:
            return None, None
        return s.min().date().isoformat(), s.max().date().isoformat()
    except Exception:
        return None, None

# ===== CSV =====
def sniff_csv(path: Path) -> dict:
    cols, rows, min_date, max_date = read_csv_preview(path)
    return {
        "kind": "csv",
        "detected": {"bank": None, "account_core_dv": None, "period_from": min_date, "period_to": max_date},
        "table": {"columns": cols, "sample": rows},
        "suggest": {"period_from": min_date, "period_to": max_date},
        "needs": {"bank": True, "account_id": False, "period_range": not (min_date and max_date)},
    }

def read_csv_preview(path: Path) -> tuple[list[str], list[list[Any]], Optional[str], Optional[str]]:
    if pd is None:
        return [], [], None, None
    try:
        df = pd.read_csv(str(path), nrows=120, sep=None, engine="python")
        cols = [str(c) for c in df.columns.tolist()]
        sample = df.head(10).fillna("").astype(object).values.tolist()
        min_d, max_d = try_parse_dates_in_df(df)
        return cols, sample, min_d, max_d
    except Exception:
        return [], [], None, None

# ===== Mapeo short → full =====
def map_short_account_to_full(short_code: str) -> Optional[str]:
    short_digits = re.sub(r"\D+", "", short_code or "")  # '26005/5' → '260055'
    if not short_digits:
        return None
    candidates: list[str] = []
    for full in ACCOUNT_MAP.keys():
        full_digits = re.sub(r"\D+", "", full)           # '3-111-0100026005-5' → '311101000260055'
        if full_digits.endswith(short_digits) or short_digits in full_digits[-12:]:
            candidates.append(full)
    if len(candidates) == 1:
        return candidates[0]
    return None
