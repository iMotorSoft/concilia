import time
import os
from pathlib import Path
import duckdb
from openpyxl import load_workbook

# Configuración de archivos para la prueba
FILES = [
    "/home/issajar/Downloads/06- Junio al 30.xlsx",
    "/home/issajar/Downloads/07- Julio al 31 (copy).xlsx"
]

def benchmark_openpyxl(file_path):
    """Simula la lógica actual de lectura 'lightweight' con openpyxl."""
    start_time = time.time()
    try:
        # Usamos read_only=True y data_only=True
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                row_count += 1
        
        wb.close()
        elapsed = time.time() - start_time
        return elapsed, row_count
    except Exception as e:
        return None, str(e)

def benchmark_duckdb(file_path, con):
    """Utiliza DuckDB con la extensión spatial para leer el Excel."""
    start_time = time.time()
    try:
        # Consultamos el número de filas ignorando nulas para que sea comparable
        res = con.execute(f"SELECT count(*) FROM st_read('{file_path}')").fetchone()
        row_count = res[0]
        
        elapsed = time.time() - start_time
        return elapsed, row_count
    except Exception as e:
        return None, str(e)

def run_benchmark():
    print("="*60)
    print(" BENCHMARK: OPENPYXL vs DUCKDB (Lectura de Excel)")
    print("="*60)
    
    # Pre-cargar DuckDB para no medir el tiempo de instalación/carga
    try:
        con = duckdb.connect(database=':memory:')
        con.execute("INSTALL spatial;")
        con.execute("LOAD spatial;")
    except Exception as e:
        print(f"Error inicializando DuckDB: {e}")
        return

    for f in FILES:
        if not os.path.exists(f):
            print(f"X Salteando: {f} (Archivo no encontrado)")
            continue
            
        print(f"\nProcesando: {os.path.basename(f)}")
        
        # openpyxl
        t_ox, c_ox = benchmark_openpyxl(f)
        if t_ox is not None:
            print(f"  [openpyxl] Filas: {c_ox} | Tiempo: {t_ox:.4f}s")
        else:
            print(f"  [openpyxl] ERROR: {c_ox}")
            
        # duckdb
        t_db, c_db = benchmark_duckdb(f, con)
        if t_db is not None:
            print(f"  [duckdb]   Filas: {c_db} | Tiempo: {t_db:.4f}s")
            if t_ox:
                speedup = t_ox / t_db
                print(f"  >>> DuckDB es {speedup:.1f}x más rápido")
        else:
            print(f"  [duckdb]   ERROR: {c_db}")

if __name__ == "__main__":
    run_benchmark()
