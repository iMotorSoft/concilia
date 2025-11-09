import openpyxl
from datetime import datetime, date, timedelta
from typing import Optional, Any

def _as_date(val: Any) -> Optional[date]:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        if isinstance(val, float) and not float(val).is_integer():
            return None
        if 20000 <= int(val) <= 80000: # Typical Excel serial date range
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=int(val))).date()
        return None
    if val not in (None, ""):
        # Try common string formats
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d', '%d-%m-%Y', '%m-%d-%Y'):
            try:
                return datetime.strptime(str(val).split(" ")[0], fmt).date()
            except ValueError:
                continue
    return None

def extract_date_range_flexible(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        min_date_found: Optional[date] = None
        max_date_found: Optional[date] = None

        # First, try to find "Fecha desde" and "Fecha hasta" in column A
        for row_idx in range(1, min(sheet.max_row + 1, 20)): # Limit search to first 20 rows
            cell_a_value = sheet.cell(row=row_idx, column=1).value # Column A
            if cell_a_value:
                if "fecha desde" in str(cell_a_value).lower():
                    date_val = sheet.cell(row=row_idx, column=2).value # Value in Column B
                    dt = _as_date(date_val)
                    if dt:
                        min_date_found = dt
                elif "fecha hasta" in str(cell_a_value).lower():
                    date_val = sheet.cell(row=row_idx, column=2).value # Value in Column B
                    dt = _as_date(date_val)
                    if dt:
                        max_date_found = dt
            
            if min_date_found and max_date_found: # If both found, we are done
                break

        if min_date_found and max_date_found:
            print(f"Rango de fechas detectado: Desde {min_date_found} hasta {max_date_found}")
            return

        # Fallback: If "Fecha desde" / "Fecha hasta" not found, try to find a column with "Fecha" in its header
        header_row_index = -1
        date_column_index = -1
        header_keyword = "Fecha"

        for row_idx in range(1, min(sheet.max_row + 1, 20)): # Limit search to first 20 rows
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value and header_keyword.lower() in str(cell_value).lower():
                    header_row_index = row_idx
                    date_column_index = col_idx
                    break
            if header_row_index != -1: # Stop searching rows if header found
                break
        
        start_data_row = 1 # Default start row for data
        if header_row_index != -1:
            start_data_row = header_row_index + 1

        if date_column_index == -1: # If still no column identified
            print(f"Error: No se pudo identificar una columna de fechas en el archivo.")
            return

        dates = []
        empty_rows_in_a_row = 0
        # Iterate through the identified column to extract dates
        for row_index in range(start_data_row, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_index, column=date_column_index).value
            dt = _as_date(cell_value)
            if dt:
                dates.append(dt)
                empty_rows_in_a_row = 0
            elif not cell_value: # Only count truly empty cells
                empty_rows_in_a_row += 1
            
            if empty_rows_in_a_row >= 20: # Stop if 20 consecutive empty rows
                break

        if dates:
            min_date = min(dates)
            max_date = max(dates)
            print(f"Rango de fechas detectado: Desde {min_date} hasta {max_date}")
        else:
            print("No se encontraron fechas válidas en la columna especificada.")

    except FileNotFoundError:
        print(f"Error: El archivo no se encontró en la ruta: {file_path}")
    except Exception as e:
        print(f"Ocurrió un error: {e}")

if __name__ == "__main__":
    excel_file_path = "/media/issajar/DEVELOP/Projects/iMotorSoft/ai/dev/SpendIQ/Doc/FCE/Conciliacion/EXTRACTO SANTANDER AGOST-SEPT.xlsx"
    extract_date_range_flexible(excel_file_path)