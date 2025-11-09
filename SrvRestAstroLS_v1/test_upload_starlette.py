# test_upload_starlette.py
from __future__ import annotations
import os, re, traceback, json
from pathlib import Path
from typing import Optional, Any, List, Tuple

from starlette.applications import Starlette
from starlette.responses import JSONResponse
from starlette.routing import Route

# === Opcional: pandas/openpyxl para Excel ===
try:
    import pandas as pd  # type: ignore
except Exception:  # pragma: no cover
    pd = None

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None


UPLOAD_DIR = Path("./_uploads").resolve()
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

SAFE_NAME_RX = re.compile(r"[^A-Za-z0-9._-]+")

def sanitize_filename(name: str) -> str:
    return (SAFE_NAME_RX.sub("_", os.path.basename(name or "upload.bin")).strip("._") or "upload.bin")


async def health(request):
    return JSONResponse({"ok": True, "upload_dir": str(UPLOAD_DIR), "pandas": bool(pd), "openpyxl": bool(load_workbook)})


async def upload(request):
    """
    POST /upload   (multipart/form-data)
      - file: UploadFile (xlsx / csv / lo que sea)

    Responde:
      { ok, filename, saved_as, bytes_written, content_type, preview: { ... } }
    """
    try:
        form = await request.form()
        file = form.get("file")
        if file is None:
            return JSONResponse({"ok": False, "error": "Falta campo 'file' en multipart."}, status_code=400)

        original_name = file.filename or "upload.bin"
        content_type  = file.content_type
        safe_name     = sanitize_filename(original_name)
        dest          = UPLOAD_DIR / safe_name

        # --- streaming por bloques ---
        bytes_written = 0
        with open(dest, "wb") as out:
            while True:
                chunk = await file.read(1024 * 1024)
                if not chunk:
                    break
                out.write(chunk)
                bytes_written += len(chunk)

        # --- vista previa: si es xlsx intentamos leer ---
        preview: dict[str, Any] = {}
        if str(dest).lower().endswith((".xlsx", ".xls")):
            preview = sniff_excel(dest)
        elif str(dest).lower().endswith(".csv"):
            preview = sniff_csv(dest)
        else:
            preview = {"kind": "unknown", "note": "Formato no reconocido (no es xlsx/xls/csv)."}

        return JSONResponse({
            "ok": True,
            "filename": original_name,
            "saved_as": str(dest),
            "bytes_written": bytes_written,
            "content_type": content_type,
            "preview": preview
        })

    except Exception as e:
        err = f"{type(e).__name__}: {e}"
        tb  = traceback.format_exc(limit=8)
        print("UPLOAD ERROR:", err, "\n", tb, flush=True)
        return JSONResponse({"ok": False, "error": err, "trace": tb}, status_code=500)


# ---------- Sniffers ----------

BANK_HINTS = [
    ("BANCO CIUDAD", "ciudad"),
    ("BANCO SANTANDER", "santander"),
    ("SANTANDER", "santander"),
    ("BANCO PATAGONIA", "patagonia"),
    ("PATAGONIA", "patagonia"),
    ("BANCO", None),  # genérico
]

RE_CC_LINE = re.compile(r"(CC\s*\$|C\/C|\bCTA\.? CTE\b|\bCUENTA\s*CORRIENTE\b)", re.IGNORECASE)
RE_ACCOUNT = re.compile(r"(\d[\d\-\./ ]{6,}\d)")  # secuencia con dígitos y separadores
RE_PERIOD  = re.compile(r"(0?[1-9]|[12][0-9]|3[01])[/\.-](0?[1-9]|1[0-2])[/\.-]([12]\d{3})")  # dd/mm/yyyy o variantes

def sniff_excel(path: Path) -> dict:
    """Vista previa básica de Excel: header lines (openpyxl) + muestra (pandas), columnas y fechas."""
    header_excerpt = read_excel_header_text(path, max_lines=6)
    bank, account_core_dv = detect_bank_account(header_excerpt)
    kind = "bank_movements" if RE_CC_LINE.search(header_excerpt or "") else "gl_or_unknown"

    cols, rows, min_date, max_date = read_table_preview(path)

    return {
        "kind": kind,
        "detected": {
            "bank": bank,
            "account_core_dv": account_core_dv,
            "header_excerpt": header_excerpt.strip() if header_excerpt else None,
        },
        "table": {
            "columns": cols,
            "sample": rows,  # primeras ~10 filas (serializable)
        },
        "suggest": {
            "period_from": min_date,
            "period_to": max_date,
        }
    }

def sniff_csv(path: Path) -> dict:
    cols, rows, min_date, max_date = read_csv_preview(path)
    return {
        "kind": "csv",
        "table": {"columns": cols, "sample": rows},
        "suggest": {"period_from": min_date, "period_to": max_date},
    }

def read_excel_header_text(path: Path, max_lines: int = 8) -> str:
    if not load_workbook:
        return ""
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        lines: List[str] = []
        # leemos primeras filas como texto “plano”
        for r in ws.iter_rows(min_row=1, max_row=max_lines, values_only=True):
            line = " ".join([str(c) for c in r if c not in (None, "")])
            if line.strip():
                lines.append(line.strip())
        wb.close()
        return "\n".join(lines[:max_lines])
    except Exception:
        return ""

def detect_bank_account(header_text: str | None) -> Tuple[Optional[str], Optional[str]]:
    if not header_text:
        return None, None
    text_up = header_text.upper()
    bank: Optional[str] = None
    for key, label in BANK_HINTS:
        if key in text_up:
            bank = label or "bank"
            break
    # cuenta “core/dv” aproximada (tomamos la 1ra secuencia larga con separadores)
    m = RE_ACCOUNT.search(header_text)
    account_core_dv = m.group(1) if m else None
    return bank, account_core_dv

def try_parse_dates_in_df(df) -> Tuple[Optional[str], Optional[str]]:
    if df is None or df.empty:
        return None, None
    # heurística: encontrar primera columna con dtype datetime o parseable
    date_col = None
    for c in df.columns:
        s = df[c]
        if str(s.dtype).startswith("datetime"):
            date_col = c
            break
        # intento parsear si es object
        if s.dtype == "object":
            try:
                parsed = pd.to_datetime(s, errors="coerce", dayfirst=True, infer_datetime_format=True)
                if parsed.notna().sum() >= max(3, int(len(parsed) * 0.1)):  # suficiente densidad
                    df[c] = parsed
                    date_col = c
                    break
            except Exception:
                pass
    if not date_col:
        return None, None
    try:
        s = pd.to_datetime(df[date_col], errors="coerce")
        s = s.dropna()
        if s.empty:
            return None, None
        return s.min().date().isoformat(), s.max().date().isoformat()
    except Exception:
        return None, None

def read_table_preview(path: Path) -> Tuple[List[str], List[List[Any]], Optional[str], Optional[str]]:
    if not pd:
        return [], [], None, None
    try:
        df = pd.read_excel(str(path), engine="openpyxl", nrows=120)
        cols = [str(c) for c in df.columns.tolist()]
        # muestras serializables
        sample = df.head(10).fillna("").astype(object).values.tolist()
        min_d, max_d = try_parse_dates_in_df(df)
        return cols, sample, min_d, max_d
    except Exception:
        # fallback sin pandas: devolvemos vacío
        return [], [], None, None

def read_csv_preview(path: Path) -> Tuple[List[str], List[List[Any]], Optional[str], Optional[str]]:
    if not pd:
        return [], [], None, None
    try:
        df = pd.read_csv(str(path), nrows=120, sep=None, engine="python")  # autodetecta separador
        cols = [str(c) for c in df.columns.tolist()]
        sample = df.head(10).fillna("").astype(object).values.tolist()
        min_d, max_d = try_parse_dates_in_df(df)
        return cols, sample, min_d, max_d
    except Exception:
        return [], [], None, None


app = Starlette(routes=[
    Route("/health", health),
    Route("/upload", upload, methods=["POST"]),
])

