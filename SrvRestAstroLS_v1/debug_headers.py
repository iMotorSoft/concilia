import openpyxl

def debug_excel_headers(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        print("Contenido de la primera fila (encabezados):")
        for col_index in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col_index).value
            print(f"Columna {col_index}: {cell_value}")

    except FileNotFoundError:
        print(f"Error: El archivo no se encontró en la ruta: {file_path}")
    except Exception as e:
        print(f"Ocurrió un error: {e}")

if __name__ == "__main__":
    excel_file_path = "/media/issajar/DEVELOP/Projects/iMotorSoft/ai/dev/SpendIQ/Doc/FCE/Conciliacion/EXTRACTO SANTANDER AGOST-SEPT.xlsx"
    debug_excel_headers(excel_file_path)
