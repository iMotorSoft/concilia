# -*- coding: utf-8 -*-
# SrvRestAstroLS_v1/routes/v1/uploads_v2_concilia.py
from __future__ import annotations
import asyncio
import shutil
import traceback
from pathlib import Path
from urllib.parse import urlparse
from uuid import uuid4
from typing import Any, Optional

from litestar import post
from litestar.response import Response
from litestar.enums import MediaType  # 👈 usamos MediaType.JSON

import globalVar as Var
from .agui_notify import emit
from services.ingest.sniff_bank import sniff_file

try:
    from openpyxl import load_workbook, Workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None
    Workbook = None

from datetime import date, timedelta
import re

def _merge_validation_for_role(intel: dict, role: str) -> dict | None:
    """Combina la validación base con un error de tipo si role != kind detectado."""
    base = intel.get("validation") or None
    kind = (intel.get("kind") or "").lower()
    mismatch_error: str | None = None
    if role == "extracto" and kind and kind != "bank_movements":
        mismatch_error = f"Se detectó tipo '{kind}' y no parece extracto bancario."
    if role == "contable" and kind and kind != "gl":
        mismatch_error = f"Se detectó tipo '{kind}' y no parece contable/PILAGA."

    if not mismatch_error:
        return base

    errors = list(base.get("errors") or []) if base else []
    warnings = list(base.get("warnings") or []) if base else []
    errors.append(mismatch_error)
    return {"is_valid": False, "errors": errors, "warnings": warnings}

def _parse_iso_date(s: Any) -> date | None:
    if not s:
        return None
    try:
        return date.fromisoformat(str(s).strip())
    except Exception:
        return None

def _iso(d: date | None) -> str | None:
    return d.isoformat() if d else None

def _month_iter(start: date, end: date) -> list[tuple[int, int]]:
    months: list[tuple[int, int]] = []
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        months.append((y, m))
        if m == 12:
            y += 1
            m = 1
        else:
            m += 1
    return months

def _month_bounds(y: int, m: int) -> tuple[date, date]:
    first = date(y, m, 1)
    if m == 12:
        next_first = date(y + 1, 1, 1)
    else:
        next_first = date(y, m + 1, 1)
    return first, next_first - timedelta(days=1)

_RE_DMY = re.compile(r"^\s*(\d{1,2})[/\.-](\d{1,2})[/\.-]([12]\d{3})\s*$")

def _try_parse_date_cell(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, date):
        return v
    try:
        import datetime as _dt
        if isinstance(v, _dt.datetime):
            return v.date()
    except Exception:
        pass

    s = str(v).strip()
    if not s:
        return None
    m = _RE_DMY.match(s)
    if m:
        dd, mm, yyyy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(yyyy, mm, dd)
        except Exception:
            return None
    try:
        import pandas as pd  # type: ignore
        dt = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.notna(dt):
            if hasattr(dt, "iloc"):
                dt = dt.iloc[0]
            return dt.date()
    except Exception:
        return None
    return None

def _extract_present_dates_from_xlsx(path: Path, max_rows: int = 60000) -> set[date]:
    """Extrae fechas distintas del XLSX (columna FECHA) para estimar cobertura por días."""
    if load_workbook is None:
        return set()
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = _pick_worksheet(wb)
        hdr = _find_header_row(ws)
        header_row = next(ws.iter_rows(min_row=hdr, max_row=hdr, values_only=True), None) or []
        fecha_idx = None
        for i, v in enumerate(header_row, start=1):
            if str(v or "").strip().upper().startswith("FECHA"):
                fecha_idx = i
                break
        if fecha_idx is None:
            wb.close()
            return set()

        out: set[date] = set()
        last_row = min(int(getattr(ws, "max_row", 0) or 0), max_rows)
        for row in ws.iter_rows(min_row=hdr + 1, max_row=last_row, min_col=fecha_idx, max_col=fecha_idx, values_only=True):
            v = row[0] if row else None
            d = _try_parse_date_cell(v)
            if d:
                out.add(d)
        wb.close()
        return out
    except Exception:
        return set()

def _partial_months_from_dates(dates_present: set[date], start: date, end: date) -> list[dict]:
    """
    Genera meses parciales basados en días presentes en filas (no en rango min/max).
    Importante: si un día no tiene movimientos, aparecerá como "faltante".
    """
    by_month: dict[tuple[int, int], set[int]] = {}
    for d in dates_present:
        by_month.setdefault((d.year, d.month), set()).add(d.day)

    out: list[dict] = []
    for y, m in _month_iter(start, end):
        ms, me = _month_bounds(y, m)
        days = by_month.get((y, m), set())
        if not days:
            continue
        total_days = (me - ms).days + 1
        missing_days = total_days - len(days)
        if missing_days <= 0:
            continue

        missing_ranges: list[dict] = []
        run_start: int | None = None
        run_end: int | None = None
        for day_num in range(1, total_days + 1):
            if day_num not in days:
                if run_start is None:
                    run_start = day_num
                run_end = day_num
            else:
                if run_start is not None and run_end is not None:
                    a = date(y, m, run_start)
                    b = date(y, m, run_end)
                    missing_ranges.append({"from": a.isoformat(), "to": b.isoformat(), "days": (b - a).days + 1})
                run_start = run_end = None
        if run_start is not None and run_end is not None:
            a = date(y, m, run_start)
            b = date(y, m, run_end)
            missing_ranges.append({"from": a.isoformat(), "to": b.isoformat(), "days": (b - a).days + 1})

        out.append({
            "month": f"{y:04d}-{m:02d}",
            "covered_days": len(days),
            "total_days": total_days,
            "missing_days": missing_days,
            "missing_ranges": missing_ranges,
        })
    return out

def _coverage_from_present_dates(
    dates_present: set[date],
    start: date,
    end: date,
    *,
    min_days_present: int = 7,
    min_missing_days_to_report_partial: int = 2,
) -> dict:
    """
    Cobertura basada en días presentes en las filas:
      - Un mes se considera "presente" solo si tiene al menos `min_days_present` días distintos.
      - Si está por debajo, lo marcamos como faltante (para evitar falsos positivos por 'mes pegado' a fin de mes).
    """
    by_month: dict[tuple[int, int], set[int]] = {}
    for d in dates_present:
        by_month.setdefault((d.year, d.month), set()).add(d.day)

    missing_months: list[str] = []
    missing_ym: list[tuple[int, int]] = []
    partial_months: list[dict] = []

    for y, m in _month_iter(start, end):
        ms, me = _month_bounds(y, m)
        total_days = (me - ms).days + 1
        days = by_month.get((y, m), set())
        covered_days = len(days)

        if covered_days < min_days_present:
            missing_months.append(f"{y:04d}-{m:02d}")
            missing_ym.append((y, m))
            continue

        missing_days = total_days - covered_days
        if missing_days <= 0:
            continue
        if missing_days < min_missing_days_to_report_partial:
            continue

        missing_ranges: list[dict] = []
        run_start: int | None = None
        run_end: int | None = None
        for day_num in range(1, total_days + 1):
            if day_num not in days:
                if run_start is None:
                    run_start = day_num
                run_end = day_num
            else:
                if run_start is not None and run_end is not None:
                    a = date(y, m, run_start)
                    b = date(y, m, run_end)
                    missing_ranges.append({"from": a.isoformat(), "to": b.isoformat(), "days": (b - a).days + 1})
                run_start = run_end = None
        if run_start is not None and run_end is not None:
            a = date(y, m, run_start)
            b = date(y, m, run_end)
            missing_ranges.append({"from": a.isoformat(), "to": b.isoformat(), "days": (b - a).days + 1})

        partial_months.append({
            "month": f"{y:04d}-{m:02d}",
            "covered_days": covered_days,
            "total_days": total_days,
            "missing_days": missing_days,
            "missing_ranges": missing_ranges,
        })

    # gaps por meses faltantes
    gaps: list[dict] = []
    if missing_ym:
        missing_ym.sort()
        start_ym = prev = missing_ym[0]
        for cur in missing_ym[1:]:
            py, pm = prev
            expected_next = (py + 1, 1) if pm == 12 else (py, pm + 1)
            if cur == expected_next:
                prev = cur
                continue
            gf, _ = _month_bounds(start_ym[0], start_ym[1])
            _, gt = _month_bounds(prev[0], prev[1])
            gaps.append({"from": gf.isoformat(), "to": gt.isoformat(), "days": (gt - gf).days + 1})
            start_ym = prev = cur
        gf, _ = _month_bounds(start_ym[0], start_ym[1])
        _, gt = _month_bounds(prev[0], prev[1])
        gaps.append({"from": gf.isoformat(), "to": gt.isoformat(), "days": (gt - gf).days + 1})

    return {
        "missing_months": missing_months,
        "gaps": gaps,
        "partial_months": partial_months,
        "day_presence_min_days": min_days_present,
        "partial_min_missing_days": min_missing_days_to_report_partial,
    }

def _merge_intervals(intervals: list[tuple[date, date]]) -> list[tuple[date, date]]:
    if not intervals:
        return []
    intervals = sorted(intervals, key=lambda t: t[0])
    merged: list[tuple[date, date]] = [intervals[0]]
    for a, b in intervals[1:]:
        ca, cb = merged[-1]
        if a <= (cb + timedelta(days=1)):
            merged[-1] = (ca, max(cb, b))
        else:
            merged.append((a, b))
    return merged

def _compute_overlaps_from_periods(saved: list[dict], *, max_examples: int = 6) -> dict:
    """
    Nivel A (UX): solapamiento por rangos detectados por archivo.
    Devuelve días únicos de solapamiento y algunos ejemplos (pares de archivos).
    """
    items: list[dict] = []
    for s in saved:
        pf = _parse_iso_date(s.get("period_from"))
        pt = _parse_iso_date(s.get("period_to"))
        if not (pf and pt):
            continue
        if pt < pf:
            pf, pt = pt, pf
        items.append({"filename": s.get("filename") or "", "from": pf, "to": pt})

    if len(items) < 2:
        return {"days_total": 0, "examples": []}

    intersections: list[tuple[date, date]] = []
    examples: list[dict] = []
    for i in range(len(items)):
        for j in range(i + 1, len(items)):
            a = max(items[i]["from"], items[j]["from"])
            b = min(items[i]["to"], items[j]["to"])
            if a <= b:
                intersections.append((a, b))
                if len(examples) < max_examples:
                    examples.append({
                        "a": items[i]["filename"],
                        "b": items[j]["filename"],
                        "from": a.isoformat(),
                        "to": b.isoformat(),
                        "days": (b - a).days + 1,
                    })

    merged = _merge_intervals(intersections)
    days_total = sum((b - a).days + 1 for a, b in merged)
    return {"days_total": days_total, "examples": examples}

def _compute_overlaps_from_present_dates(saved: list[dict], *, max_example_dates: int = 5) -> dict:
    """
    Solapamiento por fechas con movimientos (más preciso que por rangos):
    cuenta cuántas fechas (días) aparecen en 2+ archivos.
    """
    date_counts: dict[date, int] = {}
    for s in saved:
        dates = s.get("_present_dates")
        if not isinstance(dates, set):
            continue
        for d in dates:
            if isinstance(d, date):
                date_counts[d] = date_counts.get(d, 0) + 1

    overlap_dates = sorted([d for d, c in date_counts.items() if c >= 2])
    return {
        "days_total": len(overlap_dates),
        "example_dates": [d.isoformat() for d in overlap_dates[:max_example_dates]],
    }

def _compute_coverage(saved: list[dict]) -> dict:
    """
    Construye warning de cobertura usando period_from/to detectados por upload.
    - missing_months: meses sin ningún archivo (YYYY-MM)
    - gaps: tramos continuos de meses faltantes, expresados como rango de fechas (from/to/days)
    - partial_months: se completa luego, en base a días presentes en el XLSX consolidado.

    Nota: intencionalmente no reporta gaps de 1 día dentro de un mes (ej. si el primer movimiento del mes es el día 02),
    porque nuestro objetivo acá es detectar meses completos faltantes.
    """
    intervals: list[tuple[date, date]] = []
    for s in saved:
        a = _parse_iso_date(s.get("period_from"))
        b = _parse_iso_date(s.get("period_to"))
        if a and b:
            if b < a:
                a, b = b, a
            intervals.append((a, b))
    if not intervals:
        return {"missing_months": [], "gaps": []}

    overall_from = min(a for a, _ in intervals)
    overall_to = max(b for _, b in intervals)

    missing_months: list[str] = []
    missing_ym: list[tuple[int, int]] = []

    for y, m in _month_iter(overall_from, overall_to):
        # mes cubierto si algún intervalo toca el mes
        ms, me = _month_bounds(y, m)
        present = any((a <= me and b >= ms) for a, b in intervals)
        if not present:
            missing_months.append(f"{y:04d}-{m:02d}")
            missing_ym.append((y, m))

    # Agrupar meses faltantes consecutivos en un gap de fechas (por meses)
    gaps: list[dict] = []
    if missing_ym:
        missing_ym.sort()
        start = prev = missing_ym[0]
        for cur in missing_ym[1:]:
            py, pm = prev
            expected_next = (py + 1, 1) if pm == 12 else (py, pm + 1)
            if cur == expected_next:
                prev = cur
                continue
            gf, _ = _month_bounds(start[0], start[1])
            _, gt = _month_bounds(prev[0], prev[1])
            gaps.append({"from": gf.isoformat(), "to": gt.isoformat(), "days": (gt - gf).days + 1})
            start = prev = cur
        gf, _ = _month_bounds(start[0], start[1])
        _, gt = _month_bounds(prev[0], prev[1])
        gaps.append({"from": gf.isoformat(), "to": gt.isoformat(), "days": (gt - gf).days + 1})

    return {"missing_months": missing_months, "gaps": gaps, "partial_months": []}

def _get_files_from_form(form: Any, key: str = "file") -> list[Any]:
    """
    Extrae archivos repetidos en multipart (`file`, `file`, `file`...).
    Litestar/Starlette pueden exponer esto con APIs distintas; tomamos la lista más larga disponible.
    """
    candidates: list[list[Any]] = []

    # Starlette: FormData.getlist()
    if hasattr(form, "getlist"):
        try:
            candidates.append(list(form.getlist(key)))  # type: ignore[attr-defined]
        except Exception:
            pass

    # Litestar MultiDict: getall()
    if hasattr(form, "getall"):
        try:
            candidates.append(list(form.getall(key)))  # type: ignore[attr-defined]
        except Exception:
            pass

    # Litestar MultiDict: multi_items() (preserva duplicados)
    if hasattr(form, "multi_items"):
        try:
            candidates.append([v for k, v in form.multi_items() if k == key])  # type: ignore[attr-defined]
        except Exception:
            pass

    # Algunos MultiDict soportan items(multi=True)
    if hasattr(form, "items"):
        try:
            items = list(form.items(multi=True))  # type: ignore[call-arg]
            candidates.append([v for k, v in items if k == key])
        except TypeError:
            pass
        except Exception:
            pass

    files = max(candidates, key=len, default=[])
    files = [f for f in files if f is not None]
    if files:
        return files

    # Fallback dict-like
    val = form.get(key) if hasattr(form, "get") else None
    if val is None:
        return []
    if isinstance(val, (list, tuple)):
        return [f for f in val if f is not None]
    return [val]

def _pick_worksheet(wb: Any) -> Any:
    name = next((n for n in getattr(wb, "sheetnames", []) if str(n).strip().lower() == "principal"), None)
    if name:
        return wb[name]
    return wb[getattr(wb, "sheetnames", [])[0]]

def _find_header_row(ws: Any, scan_rows: int = 50) -> int:
    expected = {
        "FECHA",
        "COMPROBANTE",
        "CONCEPTO/COD.OP.",
        "CONCEPTO",
        "DETALLE",
        "DESCRIPCION",
        "DESCRIPCIÓN",
        "IMPORTE",
        "MONTO",
        "SALDO",
    }
    best_idx = 1
    best_score = -1
    max_row = min(int(getattr(ws, "max_row", 0) or 0), scan_rows)
    for r in range(1, max_row + 1):
        vals: list[str] = []
        for row in ws.iter_rows(min_row=r, max_row=r, values_only=True):
            for v in row:
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                vals.append(s.upper())
        if not vals:
            continue
        has_fecha = any(v == "FECHA" or v.startswith("FECHA") for v in vals)
        score = sum(1 for v in vals if v in expected)
        if has_fecha:
            score += 1
        if score > best_score and (has_fecha or score >= 2):
            best_idx = r
            best_score = score
            if best_score >= 4:
                break
    return best_idx

def _merge_extracto_workbooks(source_paths: list[Path], out_path: Path) -> None:
    """
    Fusiona múltiples extractos XLSX en uno solo:
    - Toma el header detectado del 1er archivo.
    - Apendea filas de datos de cada archivo (post-header).
    """
    if not source_paths:
        raise ValueError("No hay archivos para fusionar.")
    if load_workbook is None or Workbook is None:
        raise RuntimeError("openpyxl no disponible para fusionar XLSX.")

    wb0 = load_workbook(source_paths[0], read_only=True, data_only=True)
    ws0 = _pick_worksheet(wb0)
    hdr0 = _find_header_row(ws0)
    header_row = next(ws0.iter_rows(min_row=hdr0, max_row=hdr0, values_only=True), None)
    if not header_row:
        raise ValueError("No se pudo detectar header en el extracto base.")
    last_col = 0
    for i, v in enumerate(header_row, start=1):
        if v is None:
            continue
        if str(v).strip() == "":
            continue
        last_col = i
    last_col = max(last_col, 1)
    header_values = list(header_row[:last_col])
    wb0.close()

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "principal"
    out_ws.append(header_values)

    for src in source_paths:
        wb = load_workbook(src, read_only=True, data_only=True)
        ws = _pick_worksheet(wb)
        hdr = _find_header_row(ws)
        for row_vals in ws.iter_rows(min_row=hdr + 1, max_col=last_col, values_only=True):
            if not row_vals:
                continue
            if all(v is None or str(v).strip() == "" for v in row_vals):
                continue
            out_ws.append(list(row_vals))
        wb.close()

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(out_path)


async def _save_upload_to_incoming(file: Any, *, prefix: str = "upload") -> tuple[str, Path, int, str]:
    original_name = getattr(file, "filename", None) or f"{prefix}_{uuid4()}.bin"
    original_name = Path(str(original_name)).name
    stored_name = f"{uuid4()}_{original_name}"

    tmp_path = Path(f"/tmp/{uuid4()}_{stored_name}")
    bytes_written = 0
    with open(tmp_path, "wb") as out:
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            out.write(chunk)
            bytes_written += len(chunk)

    original_uri = Var.resolve_storage_uri("incoming", filename=stored_name)
    if not original_uri.startswith("file://"):
        raise RuntimeError("Storage provider no soportado.")
    dst = Path(urlparse(original_uri).path)
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.move(tmp_path, dst)
    return original_uri, dst, bytes_written, original_name


async def _handle_upload(request: Any, role_required: Optional[str] = None, path_label: str = "v2") -> Response:
    try:
        role = (role_required or (request.query_params.get("role") or "")).strip().lower()
        if role not in {"extracto", "contable"}:
            return Response(
                content={"ok": False, "message": "role inválido (use extracto|contable)"},
                media_type=MediaType.JSON,
                status_code=400,
            )

        form = await request.form()
        files = _get_files_from_form(form, "file")
        threadId = form.get("threadId")
        correlationId = form.get("correlationId")

        if not files:
            return Response(
                content={"ok": False, "message": "Falta campo 'file' en multipart."},
                media_type=MediaType.JSON,
                status_code=400,
            )

        # Guardar todos a incoming (sin perder compat con 1 archivo)
        saved: list[dict] = []
        total_bytes = 0
        for f in files:
            uri, dst, b, orig_name = await _save_upload_to_incoming(f, prefix=role)
            saved_item: dict = {"original_uri": uri, "path": str(dst), "bytes_written": b, "filename": orig_name}
            # Para extractos: sniff por archivo para mostrar rango/metadata en UI (1..N).
            if role == "extracto":
                intel_u = sniff_file(dst, filename_hint=orig_name)
                det_u = dict(intel_u.get("detected", {}) or {})
                sug_u = dict(intel_u.get("suggest", {}) or {})
                days_present = None
                try:
                    if Path(dst).suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
                        present_dates = _extract_present_dates_from_xlsx(Path(dst))
                        days_present = len(present_dates) or None
                        saved_item["_present_dates"] = present_dates
                except Exception:
                    days_present = None
                saved_item.update({
                    "kind": intel_u.get("kind"),
                    "bank": det_u.get("bank"),
                    "account_full": det_u.get("account_full") or det_u.get("account_core_dv"),
                    "header_excerpt": det_u.get("header_excerpt"),
                    "period_from": sug_u.get("period_from") or det_u.get("period_from"),
                    "period_to": sug_u.get("period_to") or det_u.get("period_to"),
                    "days_present": days_present,
                })
            saved.append(saved_item)
            total_bytes += b

        coverage = {"missing_months": [], "gaps": [], "partial_months": []}
        if role == "extracto" and len(saved) > 1:
            coverage = _compute_coverage(saved)
            overlap = _compute_overlaps_from_present_dates(saved)
            if not overlap.get("days_total"):
                overlap = _compute_overlaps_from_periods(saved)
            if overlap.get("days_total"):
                coverage["overlap"] = overlap

        # Si son varios extractos, consolidamos a un XLSX único para mantener el resto del pipeline intacto.
        if role == "extracto" and len(saved) > 1:
            src_paths = [Path(s["path"]) for s in saved]
            allowed = {".xlsx", ".xlsm", ".xltx", ".xltm"}
            bad = [p.name for p in src_paths if p.suffix.lower() not in allowed]
            if bad:
                return Response(
                    content={"ok": False, "message": f"Para consolidación se esperan solo XLSX (no: {', '.join(bad)})"},
                    media_type=MediaType.JSON,
                    status_code=400,
                )
            merged_name = f"{uuid4()}_extracto_merged.xlsx"
            merged_uri = Var.resolve_storage_uri("incoming", filename=merged_name)
            if not merged_uri.startswith("file://"):
                return Response(
                    content={"ok": False, "message": "Storage provider no soportado."},
                    media_type=MediaType.JSON,
                    status_code=500,
                )
            merged_path = Path(urlparse(merged_uri).path)
            try:
                _merge_extracto_workbooks(src_paths, merged_path)
            except Exception as e:
                return Response(
                    content={"ok": False, "message": f"No se pudo consolidar extractos: {type(e).__name__}: {e}"},
                    media_type=MediaType.JSON,
                    status_code=400,
                )
            original_uri = merged_uri
            dst = merged_path
            filename = merged_name
        else:
            # Caso 1 archivo (o contable): usamos el primero como original_uri para compat.
            original_uri = saved[0]["original_uri"]
            dst = Path(saved[0]["path"])
            filename = saved[0]["filename"]

        # Sniff de contenido (sobre el consolidado si corresponde)
        intel = sniff_file(dst, filename_hint=filename)
        source_file_id = str(uuid4())
        validation = _merge_validation_for_role(intel, role)
        detected = dict(intel.get("detected", {}) or {})
        suggest = dict(intel.get("suggest", {}) or {})

        # Para extractos múltiples: completar detectado/sugerido por consenso (para UX)
        if role == "extracto" and len(saved) > 1:
            banks = {s.get("bank") for s in saved if s.get("bank")}
            bank_consensus = next(iter(banks)) if len(banks) == 1 else None

            accounts = {s.get("account_full") for s in saved if s.get("account_full")}
            account_consensus = next(iter(accounts)) if len(accounts) == 1 else None

            header_excerpt = next((s.get("header_excerpt") for s in saved if s.get("header_excerpt")), None)
            if not header_excerpt:
                header_excerpt = f"Consolidado de {len(saved)} archivos"

            period_from = min((s.get("period_from") for s in saved if s.get("period_from")), default=None)
            period_to = max((s.get("period_to") for s in saved if s.get("period_to")), default=None)

            if bank_consensus:
                detected["bank"] = bank_consensus
            if account_consensus:
                detected["account_full"] = account_consensus
                detected.setdefault("account_core_dv", account_consensus)
            if header_excerpt:
                detected["header_excerpt"] = header_excerpt
            if period_from:
                detected["period_from"] = period_from
                suggest["period_from"] = period_from
            if period_to:
                detected["period_to"] = period_to
                suggest["period_to"] = period_to

            # Cobertura por días presentes en filas (detecta faltantes internos dentro del mes)
            try:
                d0 = _parse_iso_date(suggest.get("period_from") or detected.get("period_from"))
                d1 = _parse_iso_date(suggest.get("period_to") or detected.get("period_to"))
                if d0 and d1:
                    present_dates = _extract_present_dates_from_xlsx(dst)
                    coverage.update(_coverage_from_present_dates(present_dates, d0, d1, min_days_present=7))
            except Exception:
                pass

        # Para 1 solo extracto: también calcular cobertura por días presentes (sin umbral alto)
        if role == "extracto" and len(saved) == 1:
            try:
                d0 = _parse_iso_date(suggest.get("period_from") or detected.get("period_from"))
                d1 = _parse_iso_date(suggest.get("period_to") or detected.get("period_to"))
                if d0 and d1 and Path(dst).suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
                    present_dates = _extract_present_dates_from_xlsx(dst)
                    coverage = _coverage_from_present_dates(present_dates, d0, d1, min_days_present=1)
            except Exception:
                pass

        needs = dict(intel.get("needs", {}))
        if role == "extracto" and len(saved) > 1:
            if detected.get("bank"):
                needs["bank"] = False
            if suggest.get("period_from") and suggest.get("period_to"):
                needs["period_range"] = False
        if validation is not None and validation.get("is_valid") is False and role == "extracto":
            needs["valid_extracto"] = True
        if validation is not None and validation.get("is_valid") is False and role == "contable":
            needs["valid_contable"] = True

        # 4) Emitir preview al topic por SSE
        if threadId:
            payload = {
                "type": "INGEST_PREVIEW",
                "payload": {
                    "role": role,
                    "source_file_id": source_file_id,
                    "original_uri": original_uri,
                    "detected": {
                        "bank": detected.get("bank"),
                        "account_core_dv": detected.get("account_core_dv"),
                        "account_full": detected.get("account_full"),
                        "header_excerpt": detected.get("header_excerpt"),
                        "period_from": detected.get("period_from"),
                        "period_to": detected.get("period_to"),
                    },
                    "table": intel.get("table", {}),
                    "suggest": suggest,
                    "needs": needs,
                    "kind": intel.get("kind"),
                    "validation": validation or intel.get("validation"),
                    "meta": {
                        "bytes_written": total_bytes,
                        "filename": filename,
                        "correlationId": correlationId,
                        "path": path_label,
                        "uploads_count": len(saved),
                        "coverage": coverage,
                        "uploads": [
                            {
                                "filename": s.get("filename"),
                                "original_uri": s.get("original_uri"),
                                "bytes_written": s.get("bytes_written"),
                                "kind": s.get("kind"),
                                "bank": s.get("bank"),
                                "account_full": s.get("account_full"),
                                "period_from": s.get("period_from"),
                                "period_to": s.get("period_to"),
                                "days_present": s.get("days_present"),
                            }
                            for s in saved
                        ],
                    },
                },
            }
            asyncio.create_task(emit(threadId, payload))

        # 5) Responder ya (JSON explícito)
        return Response(
            content={
                "ok": True,
                "message": "Archivo recibido. Mostrando vista previa…",
                "source_file_id": source_file_id,
                "original_uri": original_uri,
                "kind": intel.get("kind"),
                "bytes_written": total_bytes,
                "filename": filename,
                "role": role,
                "path": path_label,
                "uploads_count": len(saved),
            },
            media_type=MediaType.JSON,
            status_code=200,
        )

    except Exception as e:
        tb = traceback.format_exc(limit=12)
        print(f"[upload_ingest_{path_label}] ERROR:", type(e).__name__, str(e), flush=True)
        print(tb, flush=True)
        try:
            form = await request.form()
            threadId = form.get("threadId")
            if threadId:
                asyncio.create_task(emit(threadId, {
                    "type": "TOAST", "level": "error",
                    "message": f"Upload error: {type(e).__name__}: {e} ({path_label})"
                }))
        except Exception:
            pass
        return Response(
            content={"ok": False, "message": f"Error interno en upload ({path_label})", "error": f"{type(e).__name__}: {e}", "trace": tb},
            media_type=MediaType.JSON,
            status_code=500,
        )


# Ruta nueva (v2) — respondemos JSON
@post("/api/uploads/v2/ingest", media_type=MediaType.JSON)
async def upload_ingest_v2(request: Any) -> Response:
    return await _handle_upload(request, role_required=None, path_label="v2")


# Alias compatible (vieja) — también JSON
@post("/api/uploads/v2/ingest", media_type=MediaType.JSON)
async def upload_ingest_alias(request: Any) -> Response:
    role = (request.query_params.get("role") or "extracto").strip().lower()
    return await _handle_upload(request, role_required=role, path_label="alias")
