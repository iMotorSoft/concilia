# -*- coding: utf-8 -*-
# SrvRestAstroLS_v1/routes/v1/reconcile_start.py

from __future__ import annotations

import asyncio
import logging
import math
import re
import time
import traceback
from datetime import timedelta
from pathlib import Path
from typing import Any, Optional, Tuple

from litestar import post
from litestar.response import Response

import pandas as pd
from openpyxl import load_workbook

from .agui_notify import emit
from globalVar import AUTO_BOOTSTRAP_TENANCY, WORKSPACE_SLUG
from services.db_pg import (
    append_event as core_append_event,
    close_run as core_close_run,
    connect_db as core_connect_db,
    create_run as core_create_run,
    get_workspace_by_slug,
)
from urllib.parse import urlparse

try:
    import polars as pl  # type: ignore
except Exception:  # pragma: no cover
    pl = None

logger = logging.getLogger(__name__)

# =========================
# Helpers (IO) + cache
# =========================

# Cache simple en memoria para evitar reparsear el mismo XLSX en la misma serie de request.
_DF_CACHE: dict[tuple, pd.DataFrame] = {}

def _preferred_engine() -> str:
    """Devuelve 'pyarrow' si está disponible (más rápido), si no openpyxl."""
    try:
        import pyarrow  # noqa: F401
        return "pyarrow"
    except Exception:
        return "openpyxl"


def _df_cache_key(kind: str, path: Path) -> tuple:
    st = path.stat()
    return (kind, str(path.resolve()), st.st_mtime_ns, st.st_size)


def _from_file_uri(uri: str) -> Path:
    """
    Convierte file://... en Path usable.
    Permite también rutas planas por compat.
    """
    if uri and uri.startswith("file://"):
        return Path(urlparse(uri).path)
    return Path(uri)


# =========================
# Loaders estandarizados
# =========================
_MONEY_ALLOWED = re.compile(r"[^\d,\.\-\(\)]")


def _parse_money_value(raw: str) -> float:
    """
    Normaliza importes con formatos mixtos:
      - 1.234,56  -> decimal coma
      - 1,234.56  -> decimal punto
      - 1234,56   -> decimal coma
      - 1234.56   -> decimal punto
    También respeta paréntesis como negativo y quita símbolos extra.
    """
    if raw is None:
        return 0.0
    if isinstance(raw, (int, float)):
        if isinstance(raw, float) and math.isnan(raw):
            return 0.0
        return float(raw)

    txt = str(raw).strip()
    if not txt:
        return 0.0

    neg = False
    if "(" in txt and ")" in txt:
        neg = True

    txt = _MONEY_ALLOWED.sub("", txt)
    txt = txt.replace("(", "").replace(")", "")
    if not txt:
        return 0.0

    if txt.startswith("-"):
        neg = True
        txt = txt[1:]
    txt = txt.replace("-", "")

    last_dot = txt.rfind(".")
    last_comma = txt.rfind(",")

    if last_dot != -1 and last_comma != -1:
        if last_comma > last_dot:
            txt = txt.replace(".", "").replace(",", ".")
        else:
            txt = txt.replace(",", "")
    elif last_comma != -1:
        txt = txt.replace(".", "").replace(",", ".")
    else:
        txt = txt.replace(",", "")

    try:
        val = float(txt) if txt else 0.0
    except ValueError:
        val = 0.0

    if neg:
        val = -abs(val)
    return val


def _clean_money(s: pd.Series) -> pd.Series:
    """Normaliza importes mezclando formatos AR/intl."""
    out = s.apply(_parse_money_value)
    return out.fillna(0.0)


def _load_pilaga(path: Path) -> pd.DataFrame:
    """
    Lee PILAGA (hojas típicas: “Resumen cuenta bancaria” o “Resumen cuenta tesorería”, si no la primera).
    Busca la fila de cabecera por la palabra “Fecha” y columnas Ingresos/Egresos/Acumulado.
    Devuelve DF estandarizado:
      ['fecha','monto','documento','ingreso_bruto','egreso_bruto','origen']
    """
    cache_key = _df_cache_key("pilaga", path)
    if cache_key in _DF_CACHE:
        return _DF_CACHE[cache_key].copy()

    if path.suffix.lower() in {".parquet", ".pq"}:
        if pl is not None:
            df = pl.read_parquet(str(path)).to_pandas()
        else:
            df = pd.read_parquet(str(path))
        if "fecha" in df.columns:
            df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
        _DF_CACHE[cache_key] = df.copy()
        return df

    engine = _preferred_engine()
    try:
        xls = pd.ExcelFile(str(path), engine=engine)
    except Exception:
        xls = pd.ExcelFile(str(path), engine="openpyxl")

    # Elegir hoja contable conocida
    sheet = next(
        (
            n for n in xls.sheet_names
            if "resumen cuenta bancaria" in str(n).strip().lower()
            or "resumen cuenta tesorer" in str(n).strip().lower()
        ),
        xls.sheet_names[0],
    )

    # Leemos sin header para poder detectar la fila con "Fecha"
    raw = pd.read_excel(xls, sheet_name=sheet, header=None)

    header_idx = None
    header_row = None
    for idx in range(min(len(raw), 40)):  # primeras filas
        row_vals = raw.iloc[idx].tolist()
        norm = [str(c).strip().upper() for c in row_vals if not (pd.isna(c) or str(c).strip() == "")]
        if any("FECHA" in c for c in norm) and (any("INGRES" in c for c in norm) or any("EGRES" in c for c in norm) or any("ACUM" in c for c in norm)):
            header_idx = idx
            header_row = row_vals
            break

    if header_idx is None:
        # fallback: primera fila como header
        header_idx = 0
        header_row = raw.iloc[0].tolist()

    def _clean_col_name(val, idx) -> str:
        if pd.isna(val):
            return f"col_{idx}"
        s = str(val).strip()
        return s if s else f"col_{idx}"

    columns = [_clean_col_name(c, i) for i, c in enumerate(header_row)]
    df = raw.iloc[header_idx + 1 :].copy()
    df.columns = columns

    # Quitar columnas completamente vacías
    df = df.dropna(axis=1, how="all")

    # Localizar columnas clave
    def _find_col(substrs):
        subs = [s.lower() for s in substrs]
        for c in df.columns:
            low = str(c).lower()
            if any(s in low for s in subs):
                return c
        return None

    fecha_col = _find_col(["fecha"])
    doc_col   = _find_col(["doc", "detalle"]) or (df.columns[1] if len(df.columns) > 1 else df.columns[0])
    ing_col   = _find_col(["ingres"])
    egr_col   = _find_col(["egres"])
    acu_col   = _find_col(["acum"])

    ingreso = _clean_money(df[ing_col]) if ing_col else pd.Series([0.0] * len(df))
    egreso  = _clean_money(df[egr_col]) if egr_col else pd.Series([0.0] * len(df))

    fechas = pd.to_datetime(df[fecha_col], dayfirst=True, errors="coerce") if fecha_col else pd.to_datetime([], errors="coerce")
    monto = ingreso - egreso

    out = pd.DataFrame({
        "fecha": fechas,
        "monto": monto,
        "documento": df[doc_col].astype(str) if doc_col in df else "",
        "ingreso_bruto": ingreso,
        "egreso_bruto": egreso,
    })

    out = out.dropna(subset=["fecha"])
    out = out[out["monto"].notna()]
    out = out[out["monto"] != 0]
    out = out.loc[:, ["fecha", "monto", "documento", "ingreso_bruto", "egreso_bruto"]].copy()
    out["origen"] = "PILAGA"
    out = out.reset_index(drop=True)
    _DF_CACHE[cache_key] = out.copy()
    return out


def _get_extracto_saldos(path: Path) -> Tuple[Optional[float], Optional[float]]:
    """Lee saldos inicial/final del extracto sin alterar el loader principal."""
    try:
        wb = load_workbook(str(path), data_only=True, read_only=True)
    except Exception:
        return (None, None)
    try:
        sheet = next(
            (n for n in wb.sheetnames if str(n).strip().lower() == "principal"),
            wb.sheetnames[0],
        )
        ws = wb[sheet]
        saldo_inicial = None
        saldo_final = None
        for row in ws.iter_rows(values_only=True):
            first = row[0]
            if isinstance(first, str) and "SALDO INICIAL" in first.upper():
                saldo_inicial = _parse_money_value(row[1])
            marker = row[8] if len(row) > 8 else None
            if isinstance(marker, str) and "SALDO FINAL" in marker.upper():
                saldo_final = _parse_money_value(row[9] if len(row) > 9 else None)
                break
        return (saldo_inicial, saldo_final)
    finally:
        wb.close()


def _get_pilaga_saldos(path: Path) -> Tuple[Optional[float], Optional[float]]:
    """Lee saldos inicial/final de PILAGA desde la primera columna de resumen."""
    try:
        wb = load_workbook(str(path), data_only=True, read_only=True)
    except Exception:
        return (None, None)
    try:
        sheet = next(
            (n for n in wb.sheetnames if str(n).strip().lower() == "resumen cuenta bancaria"),
            wb.sheetnames[0],
        )
        ws = wb[sheet]
        saldo_inicial = None
        saldo_final = None
        for row in ws.iter_rows(values_only=True):
            first = row[0]
            if not isinstance(first, str):
                continue
            txt = first.strip()
            up = txt.upper()
            if up.startswith("SALDO INICIAL"):
                saldo_inicial = _parse_money_value(txt.split(":")[-1])
            elif up.startswith("SALDO FINAL"):
                saldo_final = _parse_money_value(txt.split(":")[-1])
                if saldo_inicial is not None:
                    break
        return (saldo_inicial, saldo_final)
    finally:
        wb.close()


def _find_header_row_with_fecha(df: pd.DataFrame, scan_rows: int = 50) -> Optional[int]:
    """
    Busca la fila de encabezado en la que, además de 'Fecha', aparecen otras
    columnas esperables del extracto (Comprobante, Concepto, Importe, etc.).
    Esto evita falsos positivos en filas informativas previas al detalle.
    """
    expected = {"FECHA", "COMPROBANTE", "CONCEPTO/COD.OP.", "CONCEPTO", "DETALLE",
                "DESCRIPCION", "DESCRIPCIÓN", "IMPORTE", "MONTO", "SALDO"}
    best_idx: Optional[int] = None
    best_score = -1

    for i in range(min(len(df), scan_rows)):
        vals = [str(x).strip().upper() for x in df.iloc[i].tolist()]
        if not any(vals):
            continue
        has_fecha = any(v == "FECHA" or v.startswith("FECHA") for v in vals)
        score = sum(1 for v in vals if v in expected)
        if has_fecha:
            score += 1  # favorecemos filas que tengan FECHA explícito

        if score > best_score and (has_fecha or score >= 2):
            best_idx = i
            best_score = score
            if best_score >= 4:  # heurística: suficiente evidencia
                break

    return best_idx


def _load_extracto(path: Path) -> pd.DataFrame:
    """
    Lee EXTRACTO bancario (hoja 'principal' o primera).
    Detecta encabezado (fila con 'Fecha'), normaliza monto.
    Devuelve DF con columnas estandarizadas: ['fecha','monto','documento','origen']
    """
    cache_key = _df_cache_key("extracto", path)
    if cache_key in _DF_CACHE:
        return _DF_CACHE[cache_key].copy()

    if path.suffix.lower() in {".parquet", ".pq"}:
        if pl is not None:
            out = pl.read_parquet(str(path)).to_pandas()
        else:
            out = pd.read_parquet(str(path))
        if "fecha" in out.columns:
            out["fecha"] = pd.to_datetime(out["fecha"], errors="coerce")
        if "monto" in out.columns:
            out["monto"] = pd.to_numeric(out["monto"], errors="coerce").fillna(0.0)
        if "documento" in out.columns:
            out["documento"] = out["documento"].astype(str)
        _DF_CACHE[cache_key] = out.copy()
        return out

    engine = _preferred_engine()
    try:
        xls = pd.ExcelFile(str(path), engine=engine)
    except Exception:
        xls = pd.ExcelFile(str(path), engine="openpyxl")

    sheet = next((n for n in xls.sheet_names if str(n).strip().lower() == "principal"), xls.sheet_names[0])
    raw = pd.read_excel(xls, sheet_name=sheet, header=None)

    hdr = _find_header_row_with_fecha(raw)
    if hdr is None:
        hdr = 0

    headers = [str(x or "").strip() for x in raw.iloc[hdr].tolist()]
    df = raw.iloc[hdr + 1:].copy()
    df.columns = headers
    df = df.dropna(how="all")

    # Candidatos típicos (según tu análisis)
    fecha_col = next((c for c in df.columns if str(c).strip().upper() == "FECHA"), df.columns[0])
    # Importe suele estar en 'Unnamed: 4' o 'IMPORTE' etc. Probamos:
    cand_importe = [c for c in df.columns if str(c).strip().upper() in ("IMPORTE", "IMPORTE EN $", "MONTO")]
    importe_col = cand_importe[0] if cand_importe else (df.columns[4] if len(df.columns) > 4 else df.columns[-1])

    # Documento/descripcion (opcional; si no está, igual seguimos)
    cand_doc = [c for c in df.columns if str(c).strip().upper() in ("COMPROBANTE", "DESCRIPCIÓN", "DETALLE", "DESCRIPCION")]
    doc_col = cand_doc[0] if cand_doc else (df.columns[2] if len(df.columns) > 2 else df.columns[0])

    def _col_as_series(col_name: Any) -> pd.Series:
        col = df[col_name]
        if isinstance(col, pd.DataFrame):
            return col.iloc[:, 0]
        return col

    fecha_data = _col_as_series(fecha_col)
    doc_data = _col_as_series(doc_col)
    importe_data = _col_as_series(importe_col)

    out = pd.DataFrame({
        "fecha": pd.to_datetime(fecha_data, dayfirst=True, errors="coerce"),
        "documento": doc_data.astype(str),
        "monto": _clean_money(importe_data),
    })
    out = out.dropna(subset=["fecha"])
    out = out[out["monto"] != 0]
    out["origen"] = "EXTRACTO"
    out = out.reset_index(drop=True)
    _DF_CACHE[cache_key] = out.copy()
    return out


# =========================
# Matching (± ventana días)
# =========================
def _match_one_to_one_by_amount_and_date_window(
    df_p: pd.DataFrame,
    df_b: pd.DataFrame,
    days_window: int
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Empareja uno-a-uno por monto idéntico (redondeado a 2) y |fecha_p - fecha_b| <= days_window.
    Retorna: pairs, sobrantes_pilaga, sobrantes_banco
    """
    orig_cols_p = df_p.columns
    orig_cols_b = df_b.columns
    p = df_p.reset_index(drop=True).copy()
    b = df_b.reset_index(drop=True).copy()
    p["_row_id_p"] = p.index
    b["_row_id_b"] = b.index
    p["monto_r"] = p["monto"].round(2)
    b["monto_r"] = b["monto"].round(2)

    # Join por monto
    merged = p.merge(b, on="monto_r", suffixes=("_p", "_b"))
    # Ventana de fechas
    merged["date_diff_days"] = (merged["fecha_p"] - merged["fecha_b"]).abs().dt.days
    merged = merged[merged["date_diff_days"] <= abs(int(days_window))]

    # Greedy: quedarnos con el match más cercano por monto/fecha
    merged = merged.sort_values(["monto_r", "date_diff_days", "_row_id_p", "_row_id_b"])

    used_p: set[int] = set()
    used_b: set[int] = set()
    selected_rows = []
    for record in merged.to_dict("records"):
        row_id_p = record["_row_id_p"]
        row_id_b = record["_row_id_b"]
        if row_id_p in used_p or row_id_b in used_b:
            continue
        used_p.add(row_id_p)
        used_b.add(row_id_b)
        selected_rows.append(record)

    if selected_rows:
        merged = pd.DataFrame(selected_rows, columns=merged.columns)
    else:
        merged = merged.iloc[0:0].copy()

    matched_p = set(merged["_row_id_p"])
    matched_b = set(merged["_row_id_b"])

    sobrantes_p = p[~p["_row_id_p"].isin(matched_p)][orig_cols_p].copy()
    sobrantes_b = b[~b["_row_id_b"].isin(matched_b)][orig_cols_b].copy()
    pairs = merged.drop(columns=["_row_id_p", "_row_id_b"], errors="ignore")

    return pairs.reset_index(drop=True), sobrantes_p.reset_index(drop=True), sobrantes_b.reset_index(drop=True)


# =========================
# API Route
# =========================
_RECONCILE_STAGES = [
    {"name": "PREPARE_INPUTS", "label": "Preparando entradas", "weight": 2},
    {"name": "LOAD_EXTRACTO", "label": "Cargando extracto", "weight": 18},
    {"name": "LOAD_CONTABLE", "label": "Cargando contable", "weight": 18},
    {"name": "NORMALIZE", "label": "Normalizando", "weight": 8},
    {"name": "MATCH_1_1", "label": "Conciliando 1→1", "weight": 16},
    {"name": "SUMMARY", "label": "Resumen", "weight": 10},
    {"name": "FINALIZE", "label": "Finalizando", "weight": 2},
]


@post("/api/reconcile/start")
async def reconcile_start(request: Any) -> Response:
    """
    FORM multipart o x-www-form-urlencoded:
      - threadId (opcional): para SSE
      - uri_extracto: file://... (obligatorio)
      - uri_contable: file://... (obligatorio)
      - days_window: int (opcional, default 5)

    Emite por SSE:
      - {type:"RUN_START", ...}
      - {type:"RESULTS_READY", payload:{summary, counts}}
    """
    core_conn = None
    core_run_id = None
    workspace_id = None
    workspace_slug = WORKSPACE_SLUG
    run_id = None
    stage_started: dict[str, float] = {}
    thread_id = None

    async def _bootstrap_workspace(conn: Any, slug: str) -> str:
        if "-" in slug:
            client_slug, product_slug = slug.split("-", 1)
        else:
            client_slug = slug
            product_slug = slug

        client_id = await conn.fetchval(
            "SELECT client_id FROM core_clients WHERE slug = $1",
            client_slug,
        )
        if not client_id:
            client_id = await conn.fetchval(
                """
                INSERT INTO core_clients (slug, name)
                VALUES ($1, $2)
                ON CONFLICT (slug) DO UPDATE
                  SET name = EXCLUDED.name
                RETURNING client_id
                """,
                client_slug,
                client_slug,
            )
        if not client_id:
            raise RuntimeError(f"Bootstrap failed for core_clients slug={client_slug}")

        product_id = await conn.fetchval(
            "SELECT product_id FROM core_products WHERE slug = $1",
            product_slug,
        )
        if not product_id:
            product_id = await conn.fetchval(
                """
                INSERT INTO core_products (slug, name)
                VALUES ($1, $2)
                ON CONFLICT (slug) DO UPDATE
                  SET name = EXCLUDED.name
                RETURNING product_id
                """,
                product_slug,
                product_slug,
            )
        if not product_id:
            raise RuntimeError(f"Bootstrap failed for core_products slug={product_slug}")

        workspace_id = await conn.fetchval(
            "SELECT workspace_id FROM core_workspaces WHERE slug = $1",
            slug,
        )
        if not workspace_id:
            workspace_id = await conn.fetchval(
                """
                INSERT INTO core_workspaces (slug, name, client_id, product_id)
                VALUES ($1, $2, $3, $4)
                ON CONFLICT (slug) DO UPDATE
                  SET name = EXCLUDED.name
                RETURNING workspace_id
                """,
                slug,
                slug,
                client_id,
                product_id,
            )
        if not workspace_id:
            raise RuntimeError(f"Bootstrap failed for core_workspaces slug={slug}")

        return str(workspace_id)

    async def _emit_event(payload: dict[str, Any]) -> None:
        if thread_id:
            await emit(thread_id, payload)

    async def _emit_stage(
        stage: str,
        status: str,
        message: Optional[str] = None,
        timing_ms: Optional[int] = None,
        metrics: Optional[dict[str, Any]] = None,
    ) -> None:
        payload = {
            "type": "RECONCILE_STAGE",
            "payload": {
                "run_id": run_id,
                "stage": stage,
                "status": status,
                "message": message,
                "timing_ms": timing_ms,
                "metrics": metrics,
            },
        }
        await _emit_event(payload)
        if core_conn and core_run_id and workspace_id:
            await core_append_event(
                core_conn,
                workspace_id=workspace_id,
                run_id=core_run_id,
                type="STAGE",
                payload={
                    "stage": stage,
                    "status": status,
                    "message": message,
                    "timing_ms": timing_ms,
                    "metrics": metrics,
                },
            )

    try:
        form = await request.form()
        thread_id = form.get("threadId")
        # Campos históricos del frontend: extracto_original_uri / contable_original_uri
        # Nueva versión usa uri_extracto / uri_contable. Aceptamos ambos.
        uri_extracto = form.get("extracto_original_uri") or form.get("uri_extracto") or ""
        uri_contable = form.get("contable_original_uri") or form.get("uri_contable") or ""
        days_window = int(form.get("days_window") or 5)

        if not uri_extracto or not uri_contable:
            return Response({"ok": False, "message": "Faltan URIs: uri_extracto y uri_contable son obligatorios."}, status_code=400)

        # Core-only run initialization
        core_conn = await core_connect_db()
        try:
            workspace_id = await get_workspace_by_slug(core_conn, workspace_slug)
        except ValueError:
            if AUTO_BOOTSTRAP_TENANCY:
                workspace_id = await _bootstrap_workspace(core_conn, workspace_slug)
            else:
                return Response(
                    {"ok": False, "message": f"Workspace no existe: {workspace_slug}"},
                    status_code=400,
                )
        core_run_id = await core_create_run(
            core_conn,
            workspace_id=workspace_id,
            kind="concilia_reconcile",
            params={
                "thread_id": thread_id,
                "uri_extracto": uri_extracto,
                "uri_contable": uri_contable,
                "days_window": days_window,
            },
        )
        run_id = core_run_id

        await _emit_event({
            "type": "RUN_START",
            "payload": {
                "run_id": run_id,
                "days_window": days_window,
                "stages": _RECONCILE_STAGES,
            },
        })

        path_extracto = _from_file_uri(uri_extracto)
        path_contable = _from_file_uri(uri_contable)

        # 1) Cargar
        await _emit_stage("PREPARE_INPUTS", "start", "Validando entradas…")
        stage_started["PREPARE_INPUTS"] = time.monotonic()
        await _emit_stage(
            "PREPARE_INPUTS",
            "done",
            "Entradas listas.",
            timing_ms=int((time.monotonic() - stage_started["PREPARE_INPUTS"]) * 1000),
        )

        await _emit_stage("LOAD_CONTABLE", "start", "Cargando contable…")
        stage_started["LOAD_CONTABLE"] = time.monotonic()
        df_pilaga = _load_pilaga(path_contable)
        await _emit_stage(
            "LOAD_CONTABLE",
            "done",
            "Contable cargado.",
            timing_ms=int((time.monotonic() - stage_started["LOAD_CONTABLE"]) * 1000),
            metrics={"rows": len(df_pilaga)},
        )

        await _emit_stage("LOAD_EXTRACTO", "start", "Cargando extracto…")
        stage_started["LOAD_EXTRACTO"] = time.monotonic()
        df_banco = _load_extracto(path_extracto)
        await _emit_stage(
            "LOAD_EXTRACTO",
            "done",
            "Extracto cargado.",
            timing_ms=int((time.monotonic() - stage_started["LOAD_EXTRACTO"]) * 1000),
            metrics={"rows": len(df_banco)},
        )

        await _emit_stage("NORMALIZE", "start", "Normalizando datos…")
        stage_started["NORMALIZE"] = time.monotonic()
        await _emit_stage(
            "NORMALIZE",
            "done",
            "Normalizacion completa.",
            timing_ms=int((time.monotonic() - stage_started["NORMALIZE"]) * 1000),
        )

        # 2) Conciliar
        await _emit_stage("MATCH_1_1", "start", "Conciliando 1→1…")
        stage_started["MATCH_1_1"] = time.monotonic()
        pairs, sobrantes_p, sobrantes_b = _match_one_to_one_by_amount_and_date_window(df_pilaga, df_banco, days_window)
        await _emit_stage(
            "MATCH_1_1",
            "done",
            "Conciliacion 1→1 completa.",
            timing_ms=int((time.monotonic() - stage_started["MATCH_1_1"]) * 1000),
            metrics={"pairs": len(pairs)},
        )

        # 3) Resumen
        await _emit_stage("SUMMARY", "start", "Armando resumen…")
        stage_started["SUMMARY"] = time.monotonic()
        total_p = len(df_pilaga)
        total_b = len(df_banco)
        conc_pairs = len(pairs)
        no_en_banco = len(sobrantes_p)
        no_en_pilaga = len(sobrantes_b)

        summary = {
            "movimientos_pilaga": total_p,
            "movimientos_banco": total_b,
            "conciliados_pares": conc_pairs,
            "no_en_banco": no_en_banco,    # están en PILAGA pero no en el banco
            "no_en_pilaga": no_en_pilaga,  # están en banco pero no en PILAGA
            "days_window": days_window,
        }

        await _emit_stage(
            "SUMMARY",
            "done",
            "Resumen listo.",
            timing_ms=int((time.monotonic() - stage_started["SUMMARY"]) * 1000),
            metrics={"banco": total_b, "contable": total_p},
        )

        await _emit_stage("FINALIZE", "start", "Finalizando…")
        stage_started["FINALIZE"] = time.monotonic()
        if core_conn and core_run_id and workspace_id:
            try:
                await core_close_run(
                    core_conn,
                    workspace_id=workspace_id,
                    run_id=core_run_id,
                    status="done",
                )
            except Exception as e:
                print(f"[reconcile_start][core] close_run error: {type(e).__name__}: {e}", flush=True)
        await _emit_stage(
            "FINALIZE",
            "done",
            "Finalizacion completa.",
            timing_ms=int((time.monotonic() - stage_started["FINALIZE"]) * 1000),
        )

        await _emit_event({
            "type": "RESULTS_READY",
            "payload": {
                "run_id": run_id,
                "summary": summary,
            },
        })

        return Response({"ok": True, "summary": summary}, status_code=200)

    except Exception as e:
        tb = traceback.format_exc(limit=12)
        logger.exception("reconcile_start error")
        print("[reconcile_start] ERROR:", type(e).__name__, str(e), flush=True)
        print(tb, flush=True)
        if core_conn and core_run_id and workspace_id:
            try:
                await _emit_stage(
                    "FINALIZE",
                    "error",
                    f"{type(e).__name__}: {e}",
                )
            except Exception as stage_err:
                print(
                    f"[reconcile_start][core] error stage failed: {type(stage_err).__name__}: {stage_err}",
                    flush=True,
                )
            try:
                await core_close_run(
                    core_conn,
                    workspace_id=workspace_id,
                    run_id=core_run_id,
                    status="error",
                )
            except Exception as close_err:
                print(f"[reconcile_start][core] close_run error: {type(close_err).__name__}: {close_err}", flush=True)
        if thread_id:
            await _emit_event({
                "type": "TOAST",
                "level": "error",
                "message": f"Reconcile error: {type(e).__name__}: {e}",
            })
        return Response(
            {
                "ok": False,
                "message": "Error interno en conciliación",
                "error": f"{type(e).__name__}: {e}",
                "trace": tb,
                "where": "reconcile_start",
            },
            status_code=500,
        )
    finally:
        if core_conn:
            await core_conn.close()
